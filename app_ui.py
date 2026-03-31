import streamlit as st
import pandas as pd
import time
import os
import json
import io
import re
from faster_whisper import WhisperModel
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ---------------- PAGE CONFIG ----------------
st.set_page_config(page_title="BCIM Auditor Pro", layout="wide", page_icon="🛡️")

if 'results_history' not in st.session_state:
    st.session_state.results_history = []

# --- SCORING KEYS ---
# Added 'Deduction_Reasons' for better transparency
MASTER_KEYS = [
    "Total_Score", "Grade", "Status", "ZT_Failure", "Deduction_Reasons",
    "G_Enthusiasm_10", "G_Greeting_4", "G_FronterName_1", "G_ServiceExp_6",
    "C_Tone_2", "C_NameUsage_4", "C_NoInterrupt_6", "C_ActiveListen_7", 
    "C_PainPoints_2", "C_Empathy_4", "C_Control_2", "C_Confidence_2", 
    "C_DeadAir_3", "C_Adaptability_2", "C_Consultative_2",
    "S_Rapport_6", "S_IceBreak_4", "S_Probing_3", "S_Rebuttals_3", 
    "S_Convincing_12", "S_ContactDetails_4", "S_Networking_4",
    "P_Accuracy_2", "P_Answering_3", "P_Jargon_2"
]

# ---------------- SIDEBAR (API Keys Rotation Setup) ----------------
with st.sidebar:
    st.header("💠 System Settings")
    k1 = st.text_input("API Key 1", type="password", value="AIzaSyB53fi68IbY7jA_IR_0YWtcn3SGtUd2zEY")
    k2 = st.text_input("API Key 2", type="password")
    k3 = st.text_input("API Key 3", type="password")
    
    all_keys = [k.strip() for k in [k1, k2, k3] if k.strip()]

    if st.button("🗑️ Clear Results"):
        st.session_state.results_history = []
        st.rerun()

# ---------------- LOAD WHISPER (Optimized) ----------------
@st.cache_resource
def load_whisper():
    return WhisperModel("small", device="cpu", compute_type="int8")

# ---------------- AUTO-MODEL DISCOVERY ----------------
def get_working_model():
    try:
        available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
        for preferred in ['models/gemini-1.5-flash', 'models/gemini-1.5-pro']:
            if preferred in available_models:
                return preferred
        return available_models[0] if available_models else "models/gemini-1.5-flash"
    except:
        return "models/gemini-1.5-flash"
def split_speakers(transcript, api_keys):
    try:
        genai.configure(api_key=api_keys[0])

        model = genai.GenerativeModel("gemini-pro")

        prompt = f"""
Split the following conversation into two parts:

1. Agent speech
2. Client speech

Return ONLY JSON like:
{{
"Agent": "all agent lines",
"Client": "all client lines"
}}

Transcript:
\"\"\"{transcript}\"\"\"
"""

        response = model.generate_content(prompt)

        data = safe_json_parse(response.text)

        if data:
            return data.get("Agent", ""), data.get("Client", "")

    except:
        pass

    return "", ""
# ---------------- AI AUDIT WITH KEY ROTATION ----------------
def get_ai_audit(transcript, key_index=0):
    if key_index >= len(all_keys):
        return {"Status": "FAIL", "Total_Score": 0, "ZT_Failure": "All API Keys Expired/Invalid"}

    try:
        current_key = all_keys[key_index]
        genai.configure(api_key=current_key)
        model_name = get_working_model()
        
        # Temp 0.1 provides a balance of strictness and reasoning capability
        model = genai.GenerativeModel(
            model_name=model_name,
            generation_config={"temperature": 0.1, "response_mime_type": "application/json"}
        )
        
        # DEFINING THE DYNAMIC KEYS FOR THE PROMPT
        scoring_fields = ", ".join([f'"{k}": (score out of max)' for k in MASTER_KEYS[5:]])

        prompt = f"""
        ROLE: Senior QA Auditor. Perform a cold, evidence-based audit of the BCIM Campaign transcript below.

        INPUT TRANSCRIPT: 
        "{transcript}"

        STRICT AUDIT RULES:
        1. ZERO TOLERANCE (ZT): Immediately set Status="FAIL", Grade="F", and Total_Score=0 if the agent:
           - Is rude/unprofessional.
           - Provides false information about services.
           - Misses the required closing or networking question.
        
        2. DEDUCTION LOGIC: For every point not awarded (e.g., scoring 5/10), you MUST explain why in the 'Deduction_Reasons' field. Mention specific phrases from the transcript.

        3. ACCURACY: Do not hallucinate. If the agent didn't say it, score is 0.

        OUTPUT ONLY JSON WITH THESE EXACT KEYS:
        {{
            "Status": "PASS/FAIL",
            "Grade": "A/B/C/D/F",
            "Total_Score": (sum of all individual scores),
            "ZT_Failure": "Reason if FAIL, else 'None'",
            "Deduction_Reasons": "Bullet points explaining every mark deduction based on transcript evidence",
            "Improvement_Plan": "3 actionable steps for the agent",
            {scoring_fields}
        }}
        """

        response = model.generate_content(prompt)
        
        match = re.search(r'\{.*\}', response.text, re.DOTALL)
        if match:
            return json.loads(match.group())
        raise Exception("Format Error")

    except Exception as e:
        error_str = str(e).lower()
        if any(err in error_str for err in ["quota", "429", "api_key", "404"]):
            st.warning(f"⚠️ Key {key_index + 1} failed, rotating...")
            return get_ai_audit(transcript, key_index + 1)
        
        return {"Status": "FAIL", "Total_Score": 0, "ZT_Failure": f"Error: {str(e)}"}

# ---------------- EXCEL STYLING ----------------
def apply_excel_styling(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Style Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Style Rows & Wrap Text for Deduction/Plan
    status_idx = df.columns.get_loc("Status") + 1
    
    # Identify columns that need text wrapping
    wrap_cols = []
    for i, col in enumerate(df.columns):
        if col in ["Deduction_Reasons", "Improvement_Plan", "ZT_Failure"]:
            wrap_cols.append(i + 1)
            ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = 50

    for row in range(2, ws.max_row + 1):
        val = str(ws.cell(row=row, column=status_idx).value).upper()
        row_fill = pass_fill if "PASS" in val else fail_fill
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            # Apply wrapping to text-heavy columns
            if col in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()

# ---------------- UI ----------------
st.title("🛡️ BCIM Auditor Pro (Enterprise)")
st.markdown("### Accuracy-Focused Audit with Reasoning Logic")

files = st.file_uploader("Upload Audio Files", accept_multiple_files=True, type=['wav', 'mp3'])

if files:
    if st.button("🚀 Run Professional Audit"):
        if not all_keys:
            st.error("Please enter at least one API Key!")
            st.stop()

        whisper_model = load_whisper()
        progress_bar = st.progress(0)
        status_box = st.empty()
        table_box = st.empty()
        
        for i, file in enumerate(files):
            temp_path = f"temp_{file.name}"
            with open(temp_path, "wb") as f:
                f.write(file.getbuffer())

            try:
                status_box.info(f"🎙️ Transcribing: {file.name}")
                # High beam size for better accuracy
                segments, _ = whisper_model.transcribe(temp_path, beam_size=5, language="en")
                transcript = " ".join([s.text for s in segments])

                status_box.warning(f"🧠 Reasoning & Auditing: {file.name}")
                data = get_ai_audit(transcript, key_index=0)
                
                data["File_Name"] = file.name
                data["Date"] = time.ctime()
                st.session_state.results_history.append(data)
                
                table_box.dataframe(pd.DataFrame(st.session_state.results_history), use_container_width=True)

            except Exception as e:
                st.error(f"Critical Error on {file.name}: {e}")
            finally:
                if os.path.exists(temp_path): os.remove(temp_path)
            
            progress_bar.progress((i + 1) / len(files))

        status_box.success("✅ Batch Audit Completed!")

# ---------------- FINAL REPORT & DOWNLOAD ----------------
if st.session_state.results_history:
    st.divider()
    final_df = pd.DataFrame(st.session_state.results_history)
    
    # Reordering columns to put the new Reasoning column near the front
    cols = list(final_df.columns)
    if "Deduction_Reasons" in cols:
        cols.insert(4, cols.pop(cols.index("Deduction_Reasons")))
        final_df = final_df[cols]

    st.subheader("📊 Master Audit Report")
    st.dataframe(final_df, use_container_width=True)
    
    styled_excel = apply_excel_styling(final_df)
    st.download_button(
        label="📥 Download Professional Excel Report",
        data=styled_excel,
        file_name="BCIM_MASTER_AUDIT.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )